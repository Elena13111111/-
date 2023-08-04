import  xlsxwriter


#Импортируем необходимые модули
import openpyxl as op
file_path = 'fio.xlsx'


list_tab = [
    ["ФИО", "Должность","Подразделение"],
    ["Иванов Сергей Николаевич", "Водитель","Транспортный участок"],
    ["Петров Николай Алексеевич", "Водитель","Транспортный участок"],
    ["Пономарев Олег Витальевич", "Механик","Транспортный участок"]
    ]

excel_doc = op.Workbook()#Создаем рабочую книгу
excel_doc.create_sheet(title = 'Лист1', index = 0)#Создаем лист
sheetnames = excel_doc.sheetnames #Получение списка листов книги
sheet = excel_doc[sheetnames[0]]
i = 1
for row in list_tab:
    a,b,c = row #Распаковываем значения списка в переменные
    sheet[f"A{i}"] = a #Записываем значения в таблицу
    sheet[f"B{i}"] = b
    sheet[f"C{i}"] = c
    i += 1

excel_doc.save('fio.xlsx') #Сохраняем книгу